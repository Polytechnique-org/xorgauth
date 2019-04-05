#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Parse an Excel file from the school

Here is the procedure:
* Download a Excel export (.xlsx) from the school website
* Open it in LibreOffice
* Save in CSV format (usigned delimiter , and quote ")
* Invoke this script with the CSV file as parameter. The result can be imported
  on www.polytechnique.org admin page.

If the AX gave a file with AX IDs, load it too with -a option.

After importing the result on www.polytechnique.org, the matching between AX ID
and X.org IDs can be obtained through this SQL query (example for Bachelor 2018):

    SELECT pd.public_name, p.hrpid, p.ax_id FROM profiles AS p LEFT JOIN profile_display AS pd ON pd.pid=p.pid \
        WHERE p.hrpid LIKE '%.b2018';
"""
import argparse
import csv
import os
import os.path
import re
import sys

from openpyxl import load_workbook, Workbook


def fix_name_case(name):
    """Fix the case of some names in order to match AX file"""
    name = name.strip().title()
    if name == 'Clemence':
        name = 'Clémence'
    return name


def parse_ax_file(filepath):
    """Load a CSV file given by the AX"""
    header_row = None
    ax_data = {}
    with open(filepath, 'r') as fcsv:
        reader = csv.reader(fcsv, delimiter=';', quotechar='"', escapechar='\\', strict=True)
        for row in reader:
            if header_row is None:
                # Normalize headers
                header_row = [field_name.lower() for field_name in row]
                if 'matricule ax' not in header_row:
                    # The AX ID may be named differently
                    for possible_axid_name in ('ident ax', 'id ax'):
                        if possible_axid_name in header_row:
                            header_row[header_row.index(possible_axid_name)] = 'matricule ax'
                            break
                continue
            assert len(header_row) == len(row), "CSV file is ill-formed"""
            row_fields = dict(zip(header_row, row))
            # print(row_fields)
            first_name = fix_name_case(row_fields['prénom'])
            last_name = fix_name_case(row_fields['nom'])
            birthdate = row_fields['date de naissance'].strip()
            ax_id = row_fields['matricule ax'].strip()

            ax_data[(last_name, first_name)] = {
                'birthdate': birthdate,
                'ax_id': ax_id,
            }
    return ax_data


def main():
    parser = argparse.ArgumentParser(description="Parse a CSV file and output lists for Polytechnique.org website")
    parser.add_argument('school_excel_file', type=str,
                        help="Excel file produced by the Ecole polytechnique (Guichet_ORG_DD_MM_YYYY.xlsx)")
    parser.add_argument('-a', '--axfile', type=str,
                        help="CSV file with AX IDs")
    parser.add_argument('-o', '--outdir', type=str,
                        help="output CSV files in this directory")
    args = parser.parse_args()

    ax_data = {}
    if args.axfile:
        ax_data = parse_ax_file(args.axfile)

    data_by_promo = {}

    header_row = None
    school_data = []
    if args.school_excel_file.lower().endswith('.csv'):
        # Parse a CSV file from School's Excel file
        with open(args.school_excel_file, 'r') as fcsv:
            reader = csv.reader(fcsv, delimiter=',', quotechar='"', escapechar='\\', strict=True)
            for row in reader:
                if header_row is None:
                    header_row = [field_name.lower() for field_name in row]
                    continue
                assert len(header_row) == len(row), "CSV file is ill-formed"
                row_fields = dict(zip(header_row, row))
                school_data.append(row_fields)
    elif args.school_excel_file.lower().endswith('.xlsx'):
        # Parse School's Excel file
        school_wb = load_workbook(filename=args.school_excel_file, read_only=True)
        school_ws = school_wb.worksheets[0]
        for school_row in school_ws.iter_rows():
            if header_row is None:
                # Find the header line
                # From March 2019, the file has a new first row, with the name of the file
                # ... skip it
                if school_row[0].row == 1 and len(school_row) == 1:
                    continue
                header_row = [cell.value.lower() for cell in school_row]
                continue
            if len(school_row) < len(header_row):
                # Add enough empty fields to fill the row
                school_row += tuple([None] * (len(header_row) - len(school_row)))
            assert len(header_row) == len(school_row), "Excel file is ill-formed"
            row_fields = dict(
                (name, str(cell.value) if cell is not None and cell.value is not None else '')
                for (name, cell) in zip(header_row, school_row))
            school_data.append(row_fields)
    else:
        parser.error("Unsupported file extension of school CSV or Excel file ({})".format(
            repr(args.school_excel_file)))

    for school_fields in school_data:
        first_name = fix_name_case(school_fields['prénom'])
        last_name = fix_name_case(school_fields['nom'])
        birthdate = school_fields['date de naissance']
        sex = {'M': 'M', 'MME': 'F', '': '?'}[school_fields['civilité']]
        school_id = school_fields['matricule']
        promo_kind = school_fields['type']
        promo_year = school_fields['promotion']

        # Reject people who did not agree to share their data
        if not birthdate:
            continue

        # Graduate degrees have promotion = entry year
        if promo_kind in ('GD', 'BC'):
            promo_year = school_fields['année admission']

        # Get the AX ID
        ax_id = ''
        ax_key = (last_name, first_name)
        if ax_key in ax_data:
            row_ax_data = ax_data[ax_key]
            if row_ax_data['birthdate'] and row_ax_data['birthdate'] != birthdate:
                sys.stderr.write("Error: mismatched birthdate for {}: {} vs {}\n".format(
                    repr(ax_key),
                    row_ax_data['birthdate'],
                    birthdate,
                ))
            else:
                ax_id = row_ax_data['ax_id']
        elif ax_data:
            sys.stderr.write("Warning: {} not in AX data ({} {})\n".format(
                repr(ax_key), promo_kind, promo_year))

        # Table header:
        #   Nom Prénom Date de naissance (JJ/MM/AAAA) Sexe (F/M) Matricule École Matricule AX
        xorg_fields = (
            last_name,
            first_name,
            birthdate,
            sex,
            school_id,
            ax_id,
        )
        promo = promo_kind + ' ' + promo_year
        if promo not in data_by_promo:
            data_by_promo[promo] = []
        data_by_promo[promo].append(xorg_fields)

    if args.outdir:
        if not os.path.exists(args.outdir):
            os.makedirs(args.outdir)
        for promo, promo_data in sorted(data_by_promo.items()):
            # Sanitize the promo before using it as a file name
            if not re.match(r'^[a-zA-Z0-9 -]+$', promo):
                sys.stderr.write("Error: invalid promotion {}\n".format(repr(promo)))
                continue
            filepath = os.path.join(args.outdir, '{}.csv'.format(promo))
            print("Writing {}".format(filepath))
            with open(filepath, 'w') as fcsv:
                # X.org uses ; instead of , as delimiter
                writer = csv.writer(fcsv, delimiter=';', quotechar='"', escapechar='\\')
                writer.writerow(['Nom', 'Prénom', 'Date de naissance', 'Sexe', 'Matricule École', 'Matricule AX'])
                for xorg_fields in promo_data:
                    writer.writerow(xorg_fields)
    else:
        # Output to standard output
        for promo, promo_data in sorted(data_by_promo.items()):
            print("### {}".format(promo))
            for xorg_fields in promo_data:
                print(';'.join(xorg_fields))
            print("")


if __name__ == '__main__':
    main()
