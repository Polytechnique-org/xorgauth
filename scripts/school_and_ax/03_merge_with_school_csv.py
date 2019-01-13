#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Produce an Excel file suitable for import on AX website, from an Excel file
from the school and a JSON export of Polytechnique.org's database

Keep all the fields from the Excel file from the school and add these fields
from Polytechnique.org's database:
- AX identifier
- Polytechnique.org identifier
- Polytechnique.org email address

This requires openpyxl library:

    pip install openpyxl
"""
import argparse
import csv
import json
import re
import sys
from openpyxl import load_workbook, Workbook


# Fields in the output file, with their origin file and field name
# Multiple iterations of a field is used to verify the data
OUTPUT_FIELDS = (
    ('Matricule École', 'school', 'matricule'),
    ('Matricule AX', 'xorg', 'ax_id'),
    ('Identifiant X.org', 'xorg', 'hrid'),
    ('Couriel X.org', 'xorg', 'email'),
    ('Civilité (École)', 'school', 'civilité'),
    ('Civilité (X.org)', 'xorg', 'civility'),
    ('Nom (École)', 'school', 'nom'),
    ('Nom (X.org)', 'xorg', 'lastname'),
    ('Prénom (École)', 'school', 'prénom'),
    ('Prénom (X.org)', 'xorg', 'firstname'),
    ('Autre prénoms', 'school', 'autres prénoms'),
    ('Date de naissance', 'school', 'date de naissance'),
    ('Ville de naissance', 'school', 'ville de naissance'),
    ('Pays de naissance', 'school', 'pays de naissance'),
    ('Type de filière (École)', 'school', 'type'),
    ('Type de filière (X.org)', 'xorg', 'degree'),
    ('Promotion (École)', 'school', 'promotion'),
    ('Année d\'admission (École)', 'school', 'année admission'),
    ('Promotion (X.org)', 'xorg', 'study_year'),
    ('Date de modification (École)', 'school', 'date de modification'),
)


def main():
    parser = argparse.ArgumentParser(description="Produce an Excel file suitable for import on AX website")
    parser.add_argument('school_excel_file', type=str,
                        help="Excel file produced by the Ecole polytechnique (Guichet_ORG_DD_MM_YYYY.xlsx)")
    parser.add_argument('xorg_json_file', type=str,
                        help="exported JSON file from ./02_export_platal_profiles_to_json.py (exported_for_ax.json)")
    parser.add_argument('-o', '--output', type=str,
                        help="produce an Excel there (eg. export_pour_ax.xlsx), otherwise output CSV data")
    args = parser.parse_args()

    school_wb = load_workbook(filename=args.school_excel_file, read_only=True)
    school_ws = school_wb.worksheets[0]

    # Extract schoolid => values from X.org database
    xorg_data_from_schoolid = {}
    with open(args.xorg_json_file, 'r') as fxorg:
        xorg_data = json.load(fxorg)
    for profile_data in xorg_data['profiles']:
        school_id = profile_data['xorg_id']
        if school_id and school_id != '0':
            # Transform the school ID back to an ID given by the school
            if re.match(r'^20[0-9][0-9]0[0-9][0-9][0-9]$', school_id):
                school_id = '1' + school_id[2:4] + school_id[5:8]

            if school_id in xorg_data_from_schoolid:
                # raise RuntimeError("School ID {} is not unique in X.org database".format(profile_data['xorg_id']))
                # Unfortunately there are duplicate profiles in X.org database... warn about them
                print("Warning: school ID {} is not unique in X.org database ({} and {})".format(
                    profile_data['xorg_id'],
                    xorg_data_from_schoolid[school_id]['hrid'],
                    profile_data['hrpid']))

            # Build an email address
            if profile_data['promo'].lower().startswith('x'):
                assert profile_data['degree'] == 'Ingénieur'  # sanity check
                email = profile_data['hrpid'] + '@polytechnique.org'
            else:
                assert profile_data['degree'] != 'Ingénieur'  # sanity check
                email = profile_data['hrpid'] + '@alumni.polytechnique.org'

            civility_map = {
                'male': 'M',
                'female': 'MME',
            }
            xorg_data_from_schoolid[school_id] = {
                'hrid': profile_data['hrpid'],
                'email': email,
                'ax_id': profile_data['ax_id'],
                'firstname': profile_data['firstname_main'],
                'lastname': profile_data['lastname_main'],
                'civility': civility_map[profile_data['sex']],
                'study_year': profile_data['promo'],
                'degree': profile_data['degree'],
            }

    # Read header and verify every field is known
    school_column_for_data = {}
    for cell in next(school_ws.iter_rows()):
        school_column_for_data[cell.value] = cell.column - 1
    for field_name, origin, origin_name in OUTPUT_FIELDS:
        assert origin in ('school', 'xorg'), "Invalid origin {} in OUTPUT_FIELDS".format(repr(origin))
        if origin == 'school' and origin_name not in school_column_for_data:
            raise RuntimeError("Missing field {} (for {}) in school data".format(origin_name, field_name))

    output_data = []
    school_id_column = school_column_for_data['matricule']
    birthdate_column = school_column_for_data['date de naissance']
    for school_row in school_ws.iter_rows():
        if school_row[0].row == 1:
            continue
        school_id = str(school_row[school_id_column].value)
        xorg_row = xorg_data_from_schoolid.get(school_id)
        if xorg_row is None:
            # This is normal if the birthdate is missing from the school data
            if school_row[birthdate_column].value:
                print("Warning: missing profile for school ID {}".format(school_id))

        current_row = []
        for field in OUTPUT_FIELDS:
            if field[1] == 'school':
                current_row.append(str(school_row[school_column_for_data[field[2]]].value))
            else:
                assert field[1] == 'xorg'
                if xorg_row is not None:
                    current_row.append(str(xorg_row[field[2]]))
                else:
                    current_row.append('')
        output_data.append(current_row)

    output_data.sort()

    # Create the output file
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.append([field[0] for field in OUTPUT_FIELDS])
    for row in output_data:
        ws.append(row)

    # Freeze the header row and 3 first columns
    ws.freeze_panes = ws['D2']

    # Save the output
    if args.output:
        if args.output.lower().endswith('.csv'):
            # Allow saving in CSV format instead of Excel
            with open(args.output, 'w') as fout:
                csvstream = csv.writer(fout, delimiter=',', quotechar='"', escapechar='\\')
                for row in ws.rows:
                    csvstream.writerow([cell.value for cell in row])
        else:
            wb.save(args.output)
    else:
        csvstream = csv.writer(sys.stdout, delimiter=',', quotechar='"', escapechar='\\')
        for row in ws.rows:
            csvstream.writerow([cell.value for cell in row])


if __name__ == '__main__':
    main()
